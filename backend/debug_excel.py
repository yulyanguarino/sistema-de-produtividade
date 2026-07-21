import sys
from openpyxl import load_workbook

# Pede o caminho do arquivo ao usuário
arquivo = input("Cole o caminho do arquivo Excel: ").strip('"')

try:
    workbook = load_workbook(arquivo)
    worksheet = workbook.active

    print("\n=== HEADERS (Linha 1) ===")
    headers = [cell.value for cell in worksheet[1]]
    print(headers)

    print("\n=== PRIMEIRAS 5 LINHAS DE DADOS ===")
    for row_idx, row in enumerate(worksheet.iter_rows(min_row=2, max_row=6, values_only=False), start=2):
        print(f"\nLinha {row_idx}:")
        for col_idx, cell in enumerate(row):
            header = headers[col_idx] if col_idx < len(headers) else "?"
            valor = cell.value
            tipo = type(valor).__name__
            print(f"  {header}: {valor} (tipo: {tipo})")

except Exception as e:
    print(f"Erro: {e}")
    import traceback
    traceback.print_exc()
