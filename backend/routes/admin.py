from fastapi import APIRouter, Depends, UploadFile, File, HTTPException
from sqlalchemy.orm import Session
from sqlalchemy import text
from database import get_db
import services
from openpyxl import load_workbook
from io import BytesIO
from models import Operacao, Colaborador

router = APIRouter(prefix="/admin", tags=["admin"])


@router.get("/debug-import")
def debug_import(db: Session = Depends(get_db)):
    """Testa importação com dados hardcoded"""
    from datetime import date

    try:
        # Dados de teste hardcoded (simulando Excel)
        dados = [
            {
                "DATA": date(2026, 1, 15),
                "PEDIDO": "TEST001",
                "SEPARADOR": "TEST_SEP",
                "QTD ITENS(separador)": 5,
                "CONFERENTE": "TEST_CONF",
                "QTD ITENS(conferente)": 5
            }
        ]

        # Chamar função de import
        resultado = services.importar_operacoes_excel(db, dados)

        # Contar total após import
        total = db.query(Operacao).count()

        return {
            "status": "ok",
            "resultado_import": resultado,
            "total_operacoes_agora": total
        }
    except Exception as e:
        return {
            "status": "error",
            "erro": str(e)
        }


@router.get("/test-batch-insert")
def test_batch_insert(db: Session = Depends(get_db)):
    """Testa inserção de múltiplas operações em batch"""
    from datetime import date

    try:
        # Criar 10 operações de teste
        sep = Colaborador(nome="BATCH_SEP", ativo=True)
        conf = Colaborador(nome="BATCH_CONF", ativo=True)
        db.add(sep)
        db.add(conf)
        db.flush()

        total_antes = db.query(Operacao).count()

        # Inserir 10 operações
        for i in range(10):
            op = Operacao(
                data=date(2026, 1, 1),
                pedido=f"BATCH_{i:04d}",
                separador_id=sep.id,
                qtd_itens_separados=i+1,
                conferente_id=conf.id,
                qtd_itens_conferidos=i+1
            )
            db.add(op)

        # Fazer commit
        db.commit()

        # Verificar se foram salvos
        total_depois = db.query(Operacao).count()
        inserted = total_depois - total_antes

        return {
            "status": "ok",
            "total_antes": total_antes,
            "total_depois": total_depois,
            "inseridos_com_sucesso": inserted,
            "esperados": 10
        }
    except Exception as e:
        db.rollback()
        return {
            "status": "error",
            "erro": str(e)
        }


@router.get("/diagnose-db")
def diagnose_database(db: Session = Depends(get_db)):
    """Diagnóstico completo da conexão com banco de dados"""
    from datetime import datetime
    from database import settings, engine
    import logging
    logger = logging.getLogger(__name__)

    try:
        db_url = settings.get_database_url
        banco_tipo = "SQLite" if "sqlite" in db_url else "PostgreSQL/Neon"

        # 1. Testar conexão
        logger.info("🔍 Iniciando diagnóstico do banco...")
        db.execute(text("SELECT 1"))
        conexao_ok = True
        msg_conexao = "✅ Conexão com banco OK"

        # 2. Contar registros ANTES
        total_ops_antes = db.query(Operacao).count()
        total_cols_antes = db.query(Colaborador).count()

        # 3. Criar colaborador de teste
        nome_teste = f"DIAGNOSE_{datetime.now().timestamp()}"
        teste_col = Colaborador(nome=nome_teste, ativo=True)
        db.add(teste_col)
        db.flush()
        id_teste = teste_col.id

        # 4. Commit
        db.commit()
        logger.info(f"Commit executado. ID do teste: {id_teste}")

        # 5. Verificar se foi realmente salvo
        teste_verificado = db.query(Colaborador).filter(Colaborador.id == id_teste).first()
        foi_salvo = teste_verificado is not None

        # 6. Contar registros DEPOIS
        total_cols_depois = db.query(Colaborador).count()

        # 7. Limpar teste
        if teste_verificado:
            db.delete(teste_verificado)
            db.commit()

        return {
            "status": "ok" if foi_salvo else "falha",
            "banco_tipo": banco_tipo,
            "conexao": msg_conexao if conexao_ok else "❌ Erro na conexão",
            "teste_criacao": {
                "nome": nome_teste,
                "id": id_teste,
                "foi_salvo": foi_salvo,
                "verificado_apos_commit": teste_verificado is not None
            },
            "registros": {
                "colaboradores_antes": total_cols_antes,
                "colaboradores_depois": total_cols_depois,
                "operacoes": total_ops_antes
            },
            "pool_info": {
                "pool_size": engine.pool.size() if hasattr(engine.pool, 'size') else "N/A",
                "checked_out": engine.pool.checkedout() if hasattr(engine.pool, 'checkedout') else "N/A"
            },
            "timestamp": datetime.now().isoformat()
        }
    except Exception as e:
        import traceback
        logger.error(f"❌ Erro no diagnóstico: {str(e)}\n{traceback.format_exc()}")
        return {
            "status": "error",
            "message": f"Erro ao diagnosticar banco: {str(e)}",
            "traceback": traceback.format_exc(),
            "timestamp": datetime.now().isoformat()
        }


@router.get("/count-operacoes")
def count_operacoes(db: Session = Depends(get_db)):
    """Conta quantas operações tem no banco agora"""
    from datetime import datetime
    try:
        total = db.query(Operacao).count()
        return {
            "status": "ok",
            "total_operacoes": total,
            "timestamp": datetime.now().isoformat()
        }
    except Exception as e:
        return {
            "status": "error",
            "message": str(e)
        }


@router.get("/test-import-simples")
def test_import_simples(db: Session = Depends(get_db)):
    """Testa import com dados hardcoded - simula exatamente o upload"""
    from datetime import date
    import logging
    logger = logging.getLogger(__name__)

    try:
        # Limpar dados de teste anteriores
        db.query(Operacao).filter(Operacao.pedido.ilike("TEST-%")).delete()
        db.query(Colaborador).filter(Colaborador.nome.ilike("TESTE_%")).delete()
        db.commit()
        logger.info("Limpei dados de teste anteriores")

        # Dados de teste (como viriam do Excel)
        dados = [
            {
                "DATA": date(2026, 1, 5),
                "PEDIDO": "TEST-001",
                "SEPARADOR": "TESTE_SEP",
                "QTD ITENS(separador)": 10,
                "CONFERENTE": "TESTE_CONF",
                "QTD ITENS(conferente)": 10
            },
            {
                "DATA": date(2026, 1, 6),
                "PEDIDO": "TEST-002",
                "SEPARADOR": "TESTE_SEP_2",
                "QTD ITENS(separador)": 5,
                "CONFERENTE": "TESTE_CONF_2",
                "QTD ITENS(conferente)": 5
            }
        ]

        total_antes = db.query(Operacao).count()
        logger.info(f"Total ANTES: {total_antes}")

        # Chamar a função de import
        resultado = services.importar_operacoes_excel(db, dados)

        total_depois = db.query(Operacao).count()
        logger.info(f"Total DEPOIS: {total_depois}")
        logger.info(f"Diferença: {total_depois - total_antes}")

        return {
            "status": "ok",
            "resultado_import": resultado,
            "total_antes": total_antes,
            "total_depois": total_depois,
            "realmente_inseridos": total_depois - total_antes
        }
    except Exception as e:
        import traceback
        logger.error(f"Erro: {str(e)}\n{traceback.format_exc()}")
        return {
            "status": "error",
            "message": str(e),
            "traceback": traceback.format_exc()
        }


@router.get("/test-date-conversion")
def test_date_conversion(db: Session = Depends(get_db)):
    """Testa conversão de data ISO string"""
    from datetime import datetime

    # Simular o que vem do Excel
    data_iso_string = "2026-01-05T00:00:00"

    try:
        # Tentar converter igual o código faz
        data = datetime.fromisoformat(data_iso_string.replace('Z', '+00:00')).date()

        # Tentar inserir no banco
        sep = Colaborador(nome="TEST_SEP_DATE", ativo=True)
        conf = Colaborador(nome="TEST_CONF_DATE", ativo=True)
        db.add(sep)
        db.add(conf)
        db.flush()

        op = Operacao(
            data=data,
            pedido="TEST_DATE_001",
            separador_id=sep.id,
            qtd_itens_separados=5,
            conferente_id=conf.id,
            qtd_itens_conferidos=5
        )
        db.add(op)
        db.commit()

        # Verificar se foi salvo
        verificar = db.query(Operacao).filter(Operacao.pedido == "TEST_DATE_001").first()

        return {
            "status": "ok",
            "data_convertida": str(data),
            "tipo_data": str(type(data)),
            "foi_salvo": verificar is not None,
            "timestamp": datetime.now().isoformat()
        }
    except Exception as e:
        import traceback
        db.rollback()
        return {
            "status": "error",
            "message": str(e),
            "traceback": traceback.format_exc()
        }


@router.get("/test")
def test_endpoint(db: Session = Depends(get_db)):
    """Endpoint de teste - verifica banco de dados"""
    from datetime import datetime
    from database import settings

    try:
        # Mostrar qual banco está sendo usado
        db_url = settings.get_database_url
        banco_tipo = "SQLite" if "sqlite" in db_url else "PostgreSQL/Neon"

        # Contar quantos registros tem no banco
        total_ops = db.query(Operacao).count()
        total_cols = db.query(Colaborador).count()

        # Tentar criar um registro de teste
        teste = Colaborador(nome=f"TESTE_{datetime.now().timestamp()}", ativo=True)
        db.add(teste)
        db.commit()
        db.refresh(teste)

        # Tentar ler o registro criado
        verificar = db.query(Colaborador).filter(Colaborador.id == teste.id).first()

        return {
            "status": "ok",
            "message": "Banco respondendo",
            "banco_tipo": banco_tipo,
            "timestamp": datetime.now().isoformat(),
            "total_operacoes": total_ops,
            "total_colaboradores": total_cols,
            "teste_criado": teste.nome if teste else "FALHOU",
            "teste_lido": verificar.nome if verificar else "NÃO LIDO"
        }
    except Exception as e:
        return {
            "status": "error",
            "message": f"Erro no banco: {str(e)}",
            "timestamp": datetime.now().isoformat()
        }


@router.post("/reset-db")
def reset_database(db: Session = Depends(get_db)):
    """Limpa todas as operações e colaboradores do banco (cuidado!)"""
    import logging
    logger = logging.getLogger(__name__)
    logger.critical("🚨🚨🚨 ENDPOINT /admin/reset-db FOI CHAMADO! LIMPANDO BANCO! 🚨🚨🚨")
    try:
        services.limpar_banco_dados(db)
        logger.critical("🚨 BANCO LIMPO COM SUCESSO 🚨")
        return {"status": "ok", "message": "Banco de dados limpo com sucesso"}
    except Exception as e:
        logger.error(f"❌ Erro ao limpar: {str(e)}")
        return {"status": "error", "message": str(e)}


@router.post("/import-excel")
async def import_excel(file: UploadFile = File(...), db: Session = Depends(get_db)):
    """Importa operações de um arquivo Excel (.xlsx)"""
    import logging
    logger = logging.getLogger(__name__)
    logger.info(f"🟢🟢🟢 POST /admin/import-excel RECEBIDO! Arquivo: {file.filename} 🟢🟢🟢")
    try:
        # Validar extensão
        if not file.filename.endswith(".xlsx"):
            raise HTTPException(status_code=400, detail="Arquivo deve ser .xlsx")

        # Contar ANTES
        total_antes = db.query(Operacao).count()
        logger.info(f"Total de operações ANTES do import: {total_antes}")

        # Ler arquivo
        contents = await file.read()
        workbook = load_workbook(BytesIO(contents))
        worksheet = workbook.active

        # Extrair dados
        headers = [cell.value for cell in worksheet[1]]
        dados = []

        for row in worksheet.iter_rows(min_row=2, values_only=False):
            row_dict = {}
            for col_idx, cell in enumerate(row):
                header = headers[col_idx]
                if header:
                    row_dict[header] = cell.value

            # Pular linhas vazias
            if any(v is not None for v in row_dict.values()):
                dados.append(row_dict)

        if not dados:
            raise HTTPException(status_code=400, detail="Arquivo vazio ou sem dados")

        logger.info(f"Total de linhas no arquivo: {len(dados)}")

        # Importar
        resultado = services.importar_operacoes_excel(db, dados)

        # Contar DEPOIS
        total_depois = db.query(Operacao).count()
        logger.info(f"Total de operações DEPOIS do import: {total_depois}")
        logger.info(f"Diferença: {total_depois - total_antes}")

        return {
            "status": "ok",
            "importados": resultado["importados"],
            "erros": resultado["erros"],
            "total_erros": resultado["total_erros"],
            "total_antes": total_antes,
            "total_depois": total_depois,
            "realmente_inseridos": total_depois - total_antes,
            "headers_recebidos": headers,
            "primeiro_registro": dados[0] if dados else None,
        }

    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Erro ao processar arquivo: {str(e)}")
